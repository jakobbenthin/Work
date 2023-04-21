from openpyxl import Workbook, load_workbook

file = "C:/temp/pytest.xlsx"

wb = load_workbook(file)
ws = wb.active
print(ws['A1'].value)

wb.save(file)